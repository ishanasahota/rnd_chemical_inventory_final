import qrcode
import os
import uuid
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook, Workbook

EXCEL_FILE = "attempt2.xlsx"
GENERATED_SHEET = "GeneratedBarcodes"

# Ensure Excel file and GeneratedBarcodes sheet exist on the computer
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = GENERATED_SHEET
    ws.append(["Barcode ID"])
    wb.save(EXCEL_FILE)

wb = load_workbook(EXCEL_FILE)
if GENERATED_SHEET not in wb.sheetnames:
    ws = wb.create_sheet(GENERATED_SHEET)
    ws.append(["Barcode ID"])
    wb.save(EXCEL_FILE)

ws = wb[GENERATED_SHEET]
existing_ids = {str(row[0].value).strip() for row in ws.iter_rows(min_row=2) if row[0].value}

# Folder to store QR codes
output_folder = "barcode_images"
os.makedirs(output_folder, exist_ok=True)


# Generate 105 unique barcodes! (i think it fits on a page). also before you generate more,
#clear the contents of the barcode_images folder to make sure there's no repeat barcodes
new_ids = []
for _ in range(105):
    while True:
        barcode_id = str(uuid.uuid4())[:8]  # Short unique ID
        if barcode_id not in existing_ids:
            break

    existing_ids.add(barcode_id)
    new_ids.append(barcode_id)

    # Barcode contains only the ID
    img = qrcode.make(barcode_id)

    # Save image with ID as filename
    filename = f"{barcode_id}.png"
    img.save(os.path.join(output_folder, filename))

# Append new barcode IDs to the sheet
for barcode_id in new_ids:
    ws.append([barcode_id])

wb.save(EXCEL_FILE)

print(f"Generated {len(new_ids)} new barcodes in '{output_folder}' folder.")




# === Create printable QR sheet ===
input_folder = "barcode_images"
output_file = "barcode_sheet.pdf"

# Change these if the sizing isn't right, but this works for the stickers i was told to use
dpi = 300
page_width_in, page_height_in = 8.5, 11  # Letter size
margin_in = 0.5  # 1/2 inch margin
left_margin = 150   # shift right (increase)



cols, rows = 3, 7


page_width = int(page_width_in * dpi)
page_height = int(page_height_in * dpi)
margin = int(margin_in * dpi)
margin_y = int(0.145 * dpi)

usable_width = page_width - 2 * margin
usable_height = page_height - 2 * margin

label_width = usable_width // cols
label_height = usable_height // rows

files = [f for f in os.listdir(input_folder) if f.endswith(".png")]
files.sort()

pages = []
for i in range(0, len(files), cols * rows):
    chunk = files[i:i + cols * rows]
    page = Image.new("RGB", (page_width, page_height), "white")


    for idx, filename in enumerate(chunk):
        col = idx % cols
        row = idx // cols

        x = left_margin + col * (label_width + margin)
        y = margin + row * (label_height + margin_y)


        img = Image.open(os.path.join(input_folder, filename))
        img.thumbnail((label_width, label_height), Image.Resampling.LANCZOS)

        page.paste(img, (x, y))

        # Draw label
        draw = ImageDraw.Draw(page)
        label_text = filename.replace(".png", "")
        label_font = ImageFont.load_default()
        bbox = draw.textbbox((0, 0), label_text, font=label_font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]

        draw.text((x + (label_width - text_width) // 2, y + label_height - 25), label_text, fill="black", font=label_font)

    pages.append(page)

# Save all pages to PDF
if pages:
    pages[0].save(output_file, save_all=True, append_images=pages[1:], resolution=dpi)

print(f"Saved printable PDF to: {output_file}")
