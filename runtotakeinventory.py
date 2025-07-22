# === barcodetoexcel.py ===

import tkinter as tk
from tkinter import simpledialog, messagebox
from openpyxl import load_workbook, Workbook
import os

EXCEL_FILE = "attempt2.xlsx"
INVENTORY_SHEET = "InventoryData"

# make sure the Excel file and InventoryData sheet exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = INVENTORY_SHEET
    ws.append(["Barcode ID", "Chemical Name", "Lot Number", "Bottle Nominal Volume (L)", "Remaining Quantity", "Manufacturer", "Expiry Date"])
    wb.save(EXCEL_FILE)

wb = load_workbook(EXCEL_FILE)
if INVENTORY_SHEET not in wb.sheetnames:
    ws = wb.create_sheet(INVENTORY_SHEET)
    ws.append(["Barcode ID", "Chemical Name", "Lot Number", "Bottle Nominal Volume (L)", "Remaining Quantity", "Manufacturer", "Expiry Date"])
    wb.save(EXCEL_FILE)

def prompt_for_info(barcode_id):
    # Prompt the user for each field, exit early if cancelled
    chem_name = simpledialog.askstring("Input", "Chemical Name:")
    if chem_name is None:
        return None

    lot = simpledialog.askstring("Input", "Lot Number:")
    if lot is None:
        return None

    volume = simpledialog.askfloat("Input", "Bottle Nominal Volume (L):")
    if volume is None:
        return None

    manufacturer = simpledialog.askstring("Input", "Manufacturer:")
    if manufacturer is None:
        return None

    expiry = simpledialog.askstring("Input", "Expiry Date (Write Out M DD, YYYY):")
    if expiry is None:
        return None

    return {
        "barcode_id": barcode_id,
        "chemical_name": chem_name,
        "lot": lot,
        "volume": volume,
        "remaining": volume,
        "manufacturer": manufacturer,
        "expiry": expiry
    }

def update_inventory(barcode_id):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[INVENTORY_SHEET]

    headers = {cell.value: idx for idx, cell in enumerate(ws[1])}
    id_col = headers.get("Barcode ID")

    # Look for the barcode in Excel
    for row in ws.iter_rows(min_row=2):
        if str(row[id_col].value).strip() == barcode_id:
            # If there is an existing barcode, ask how much was used
            qty_col = headers["Remaining Quantity"]
            try:
                current_qty = float(row[qty_col].value)
            except:
                current_qty = 0

            used = simpledialog.askfloat("Usage", f"How much did you use?\nCurrent: {current_qty} L")
            if used is not None:
                row[qty_col].value = max(0, current_qty - used)
                wb.save(EXCEL_FILE)
            return

    # Barcode not found — prompt for full info
    info = prompt_for_info(barcode_id)
    if info is None:
        messagebox.showinfo("Cancelled", "Entry cancelled, returning to scan screen.")
        return

    new_row = [
        info["barcode_id"],
        info["chemical_name"],
        info["lot"],
        info["volume"],
        info["remaining"],
        info["manufacturer"],
        info["expiry"]
    ]
    ws.append(new_row)
    wb.save(EXCEL_FILE)

    messagebox.showinfo("New Bottle", f"New bottle info saved for barcode {barcode_id}")

def scan_loop():
    root = tk.Tk()
    root.withdraw()

    while True:
        barcode_id = simpledialog.askstring("Scan", "Scan barcode or enter ID:")
        if not barcode_id:
            if messagebox.askyesno("Quit", "Exit scanning?"):
                break
            continue

        update_inventory(barcode_id)

if __name__ == "__main__":
    scan_loop()
