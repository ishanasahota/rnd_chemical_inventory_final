from openpyxl import load_workbook
import pymsteams

# Load workbook and sheet
wb = load_workbook('/Users/ishanasahota/Desktop/fluidAI/rnd_chemical_inventory.xlsx', data_only=True)
ws = wb['ReorderUpdates']

order_more_items = []

for row in ws.iter_rows(min_row=2, values_only=True):
    item_name = row[0]           # Column A: Chemical Name
    quantity_alert_e = row[4]    # Column E: Expiry alert column
    quantity_alert_f = row[5]    # Column F: Quantity alert column
    
    # Check if "Order More" is in either column E or F
    if quantity_alert_e == "Order More" or quantity_alert_f == "Order More":
        order_more_items.append(item_name)

# Create the item list string
item_list_str = "\n".join(f"- {item}" for item in order_more_items)

# Send Teams message
myTeamsMessage = pymsteams.connectorcard("https://nervtechnology2.webhook.office.com/webhookb2/9b48b976-ff78-47d9-84af-0c86adb460bd@c943a20e-cde2-4f23-b31e-66242a699eb2/IncomingWebhook/2efe91b202c844cd9e3ad8f664443bb3/756d53aa-a382-47d3-9d8c-2d22fa02b8a3/V24Osd4C7KfMwor8lyRXSbBRdZcuzt6wgDMRlLNuh9QtE1")

myTeamsMessage.text(
    f"**Reorder Alert:** The following chemicals need to be reordered as there is less than 20% left:\n\n{item_list_str or 'No reorder alerts found.'}"
)

myTeamsMessage.send()
