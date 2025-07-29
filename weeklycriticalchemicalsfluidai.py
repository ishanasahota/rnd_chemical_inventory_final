from openpyxl import load_workbook 
import pymsteams
from collections import defaultdict

try:
    # Load workbook and InventoryData
    wb = load_workbook('/Users/ishanasahota/Desktop/fluidAI/rnd_chemical_inventory.xlsx', data_only=True)
    ws = wb['InventoryData']
    
    #PART 1: Weekly Summary of Critical Fluids 
    keywords = [
        "buffer solution ph 4", "buffer solution ph 5", "buffer solution ph 7",
        "buffer solution ph 8", "buffer solution ph 9", "buffer solution ph 10", 
        "ph 4 buffer solution", "ph 5 buffer solution", "ph 7 buffer solution",  # Added missing comma
        "ph 8 buffer solution", "ph 9 buffer solution", "ph 10 buffer solution",  # Added missing comma
        "conductivity standard 5 ms/cm", "conductivity standard 12.88 ms/cm",
        "conductivity standard 30 ms/cm", "conductivity standard 80 ms/cm",
        "5 ms/cm conductivity standard", "12.88 ms/cm conductivity standard",
        "30 ms/cm conductivity standard", "80 ms/cm conductivity standard"
    ]
    keywords = [k.lower() for k in keywords]  # lowercase for matching
    
    critical_summary = defaultdict(float)
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[1] or "").strip().lower()  # Column B: Chemical Name (index 1)
        remaining_qty = row[4]                    # Column E: Remaining Quantity (index 4)
        
        if any(keyword in name for keyword in keywords):
            try:
                critical_summary[name] += float(remaining_qty)
            except (TypeError, ValueError):
                print(f"⚠Skipped '{name}' due to invalid Remaining Quantity: '{remaining_qty}'")
                continue
    
    # Format critical fluids message
    critical_list_lines = []
    for name, total_amount in critical_summary.items():
        display_name = name.title()  # Capitalize for output
        main_line = f"- {display_name} ➜ {total_amount:.2f} L"
        critical_list_lines.append(main_line)
        if total_amount < 6:
            warning_line = "  -Warning: Almost at 1 bottle. Please reorder immediately."
            critical_list_lines.append(warning_line)
    
    critical_list_str = "\n".join(critical_list_lines)
    
    # Send to Teams
    msg_critical = pymsteams.connectorcard("https://nervtechnology2.webhook.office.com/webhookb2/9b48b976-ff78-47d9-84af-0c86adb460bd@c943a20e-cde2-4f23-b31e-66242a699eb2/IncomingWebhook/2efe91b202c844cd9e3ad8f664443bb3/756d53aa-a382-47d3-9d8c-2d22fa02b8a3/V24Osd4C7KfMwor8lyRXSbBRdZcuzt6wgDMRlLNuh9QtE1")
    msg_critical.text(
        "**Weekly Critical Chemical Update Alert**\n\n"
        "Combined remaining amounts by fluid name:\n\n"
        f"{critical_list_str or 'No critical fluids found.'}"
    )
    msg_critical.send()
    
    # === PART 2: Expired Chemicals ===
    # Use ReorderUpdates sheet for consistency with your other scripts
    ws_reorder = wb['ReorderUpdates']
    expired_items = []
    
    for row in ws_reorder.iter_rows(min_row=2, values_only=True):
        name = str(row[0] or "").strip()  # Column A: Chemical Name
        expiry_alert = row[4]  # Column E: Expiry Alert (assuming this is the right column)
        
        if str(expiry_alert).strip() == "Expired":
            expired_items.append(name)
    
    expired_list_str = "\n".join(f"- {name}" for name in expired_items)
    
    msg_expired = pymsteams.connectorcard("https://nervtechnology2.webhook.office.com/webhookb2/9b48b976-ff78-47d9-84af-0c86adb460bd@c943a20e-cde2-4f23-b31e-66242a699eb2/IncomingWebhook/2efe91b202c844cd9e3ad8f664443bb3/756d53aa-a382-47d3-9d8c-2d22fa02b8a3/V24Osd4C7KfMwor8lyRXSbBRdZcuzt6wgDMRlLNuh9QtE1")
    msg_expired.text(
        "**Weekly Expired Chemical Alert**\n\n"
        f"The following fluids are marked as **Expired**:\n\n{expired_list_str or 'No expired items found.'}"
    )
    msg_expired.send()
    
    print(f"Weekly alerts sent successfully. Found {len(critical_summary)} critical fluids and {len(expired_items)} expired items.")

except Exception as e:
    print(f"Error sending weekly alerts: {e}")
